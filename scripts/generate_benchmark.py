"""
generate_benchmark.py — Générateur unique du Benchmark IA Souveraine (v2.0)

Lit  : data/_source_data.json   (données brutes éditables — la seule source à mettre à jour)
Écrit: data/benchmark.json      (source de vérité du site interactif)
       data/benchmark_llm_fusion_finale.xlsx  (classeur lisible)

Tous les scores dérivés (IRN pondéré, performance normalisée, coût TCO,
scores fusionnés par profil d'application) sont CALCULÉS ici — jamais saisis à la main.

Usage :
    pip install openpyxl
    python scripts/generate_benchmark.py

Refonte v2.0 :
  - 7 profils d'application (A1 Défense classifiée → A7 Collectivités)
  - matrice de décision 7×3 (type d'application × mode de déploiement)
  - panel rafraîchi + élargi (juin 2026)
"""

import json
import argparse
from pathlib import Path

ROOT = Path(__file__).parent.parent
DEFAULT_SOURCE = ROOT / "data" / "_source_data.json"
DEFAULT_JSON   = ROOT / "data" / "benchmark.json"
DEFAULT_XLSX   = ROOT / "data" / "benchmark_llm_fusion_finale.xlsx"


# ----------------------------------------------------------------------------- calculs
def compute_irn(scores, criteria):
    """IRN /5 pondéré = Σ(score_i × poids_i) / Σ(poids_i)."""
    wsum = sum(c["weight"] for c in criteria)
    num = sum(s * c["weight"] for s, c in zip(scores, criteria))
    return num / wsum


def compute_perf(cfg, leaders, w):
    """Composite multi-bench normalisé par le leader courant, ramené sur /5."""
    a = cfg["arena"]  / leaders["arena"]  * 100
    c = cfg["coding"] / leaders["coding"] * 100
    m = cfg["mmlu"]   / leaders["mmlu"]   * 100
    i = cfg["aaii"]   / leaders["aaii"]   * 100
    composite = a * w["arena"] + c * w["coding"] + m * w["mmlu"] + i * w["aaii"]
    return composite / 100 * 5, {"arena": a, "coding": c, "mmlu": m, "aaii": i, "composite": composite}


def compute_cost(cfg, tco):
    """Managé → prix listé. Self-hosted → TCO amorti au taux d'utilisation."""
    if cfg["mode_expl"] == "self_hosted":
        rate = tco["h100_per_hour"] if cfg.get("gpu_type") == "H100" else tco["a100_per_hour"]
        hourly = cfg["gpu_count"] * rate * tco["overhead"]
        tokens_h = cfg["throughput"] * 3600 * tco["utilisation"]
        return hourly / tokens_h * 1_000_000
    return cfg.get("cost_listed")


def fused(w_irn, w_perf, irn, perf):
    return w_irn * irn + w_perf * perf


# ----------------------------------------------------------------------------- JSON
def build(source):
    crit     = source["irn_criteria"]
    leaders  = source["leaders_perf"]
    pw       = source["perf_weights"]
    tco      = source["hypotheses_tco"]
    apps     = source["app_profiles"]
    legacy   = source["legacy_profiles"]

    configs = []
    for cfg in source["configurations"]:
        irn = compute_irn(cfg["scores_irn"], crit)
        perf, pct = compute_perf(cfg, leaders, pw)
        cost = compute_cost(cfg, tco)

        out = {
            "id": cfg["id"], "name": cfg["name"], "short": cfg["short"],
            "bloc": cfg["bloc"], "country": cfg["country"], "mode": cfg["mode"],
            "model": cfg["model"], "tier": cfg["tier"], "deploy": cfg["deploy"],
            "irn": irn, "scores_irn": cfg["scores_irn"], "perf": perf,
            "arena": cfg["arena"], "coding": cfg["coding"], "mmlu": cfg["mmlu"], "aaii": cfg["aaii"],
            "perf_pct": pct, "cost": cost, "throughput": cfg["throughput"], "latency": cfg["latency"],
            "mode_expl": cfg["mode_expl"], "gpu_type": cfg.get("gpu_type"),
            "gpu_count": cfg.get("gpu_count"), "perf_flag": cfg.get("perf_flag", ""),
        }
        # scores fusionnés — 7 profils d'application
        app_scores = {a["id"]: fused(a["w_irn"], a["w_perf"], irn, perf) for a in apps}
        out["app_scores"] = app_scores
        for aid, val in app_scores.items():
            out[aid] = val
        # rétro-compat : 4 profils historiques P1–P4
        for pid, p in legacy.items():
            out[pid] = fused(p["w_irn"], p["w_perf"], irn, perf)
        configs.append(out)

    deployment_matrix = [
        {"app": a["id"], "app_name": a["name"], "reco": a["reco"]} for a in apps
    ]

    result = {
        "meta": source["meta"],
        "configurations": configs,
        "app_profiles": apps,
        "deployment_modes": source["deployment_modes"],
        "deployment_matrix": deployment_matrix,
        "profiles": legacy,              # rétro-compatibilité site v1
        "irn_criteria": crit,
        "leaders_perf": leaders,
        "hypotheses_tco": tco,
    }
    return result


# ----------------------------------------------------------------------------- XLSX
def write_xlsx(data, source, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    H = Font(bold=True, color="FFFFFF")
    HF = PatternFill("solid", fgColor="4B2E83")     # violet Theodo
    TITLE = Font(bold=True, size=13, color="4B2E83")
    crit = source["irn_criteria"]
    apps = source["app_profiles"]

    def header(ws, row, cols, start=1):
        for j, c in enumerate(cols, start):
            cell = ws.cell(row=row, column=j, value=c)
            cell.font = H; cell.fill = HF; cell.alignment = Alignment(wrap_text=True, vertical="center")

    # 00_Synthese
    ws = wb.active; ws.title = "00_Synthese"
    ws["A1"] = "Benchmark LLM — Souveraineté/IRN × Performance × Exploitation"; ws["A1"].font = TITLE
    ws["A2"] = f"Synthèse v{source['meta']['version']} — {source['meta']['date']} · 7 profils d'application"
    header(ws, 4, ["Configuration", "Tier", "Bloc géo", "Mode", "IRN /5", "Perf /5", "$/1M tok", "Débit (tok/s)"] + [a["id"] for a in apps])
    r = 5
    for c in data["configurations"]:
        vals = [c["name"], c["tier"], c["bloc"], c["mode"], round(c["irn"],2), round(c["perf"],2),
                (round(c["cost"],2) if c["cost"] is not None else "—"), c["throughput"]] + [round(c[a["id"]],2) for a in apps]
        for j, v in enumerate(vals, 1): ws.cell(row=r, column=j, value=v)
        r += 1
    ws.column_dimensions["A"].width = 38

    # 01_Configurations
    ws = wb.create_sheet("01_Configurations")
    header(ws, 1, ["ID", "Configuration", "Bloc géo", "Pays", "Mode de déploiement", "Mode (image)", "Modèle repère", "Tier"])
    for i, c in enumerate(data["configurations"], 2):
        for j, v in enumerate([c["id"], c["name"], c["bloc"], c["country"], c["mode"], c["deploy"], c["model"], c["tier"]], 1):
            ws.cell(row=i, column=j, value=v)
    ws.column_dimensions["B"].width = 40; ws.column_dimensions["G"].width = 36

    # 02_IRN_Detail
    ws = wb.create_sheet("02_IRN_Detail")
    header(ws, 1, ["ID", "Configuration"] + [c["id"] for c in crit] + ["IRN /5 pondéré"])
    ws.append(["Poids", ""] + [c["weight"] for c in crit] + [""])
    for c in data["configurations"]:
        ws.append([c["id"], c["name"]] + c["scores_irn"] + [round(c["irn"], 3)])
    ws.column_dimensions["B"].width = 40

    # 03_Performance
    ws = wb.create_sheet("03_Performance")
    ws["A1"] = "Leaders : Arena {arena} · Coding {coding} · MMLU-Pro {mmlu} · AAII {aaii}".format(**data["leaders_perf"])
    header(ws, 3, ["ID", "Configuration", "Tier", "Arena", "Coding", "MMLU-Pro", "AAII", "Composite %", "Perf /5", "Hypothèse ?"])
    r = 4
    for c in data["configurations"]:
        ws.cell(row=r, column=1, value=c["id"]); ws.cell(row=r, column=2, value=c["name"])
        for j, v in enumerate([c["tier"], c["arena"], c["coding"], c["mmlu"], c["aaii"],
                               round(c["perf_pct"]["composite"],1), round(c["perf"],3), c["perf_flag"]], 3):
            ws.cell(row=r, column=j, value=v)
        r += 1
    ws.column_dimensions["B"].width = 40; ws.column_dimensions["J"].width = 28

    # 04_Exploitation
    ws = wb.create_sheet("04_Exploitation")
    ws["A1"] = "Hypothèses TCO : H100 ${h100_per_hour}/h · overhead ×{overhead} · utilisation {utilisation}".format(**data["hypotheses_tco"])
    header(ws, 3, ["ID", "Configuration", "Tier", "Mode", "GPU", "Nb GPU", "Débit (tok/s)", "Latence (s)", "$/1M tok retenu"])
    r = 4
    for c in data["configurations"]:
        for j, v in enumerate([c["id"], c["name"], c["tier"], c["mode_expl"], c.get("gpu_type") or "—",
                               c.get("gpu_count") or "—", c["throughput"], c["latency"],
                               round(c["cost"],2) if c["cost"] is not None else "—"], 1):
            ws.cell(row=r, column=j, value=v)
        r += 1
    ws.column_dimensions["B"].width = 40

    # 05_Profils_Application
    ws = wb.create_sheet("05_Profils_Application")
    ws["A1"] = "7 profils d'application — pondération Souveraineté (IRN) ↔ Performance"; ws["A1"].font = TITLE
    header(ws, 3, ["ID", "Type d'application", "Détail", "Poids IRN", "Poids Perf",
                   "SaaS public", "SaaS souverain SecNumCloud", "Auto-hébergé contrôlé"])
    r = 4
    for a in apps:
        for j, v in enumerate([a["id"], a["name"], a["subtitle"], a["w_irn"], a["w_perf"],
                               a["reco"]["public"], a["reco"]["souverain"], a["reco"]["autoheberge"]], 1):
            ws.cell(row=r, column=j, value=v)
        r += 1
    ws.column_dimensions["B"].width = 34; ws.column_dimensions["C"].width = 30
    for col in ("F", "G", "H"): ws.column_dimensions[col].width = 24

    # 06_Score_Fusionne (par profil d'application)
    ws = wb.create_sheet("06_Score_Fusionne")
    header(ws, 1, ["ID", "Configuration", "Tier", "IRN /5", "Perf /5"] + [a["id"] for a in apps])
    for c in data["configurations"]:
        ws.append([c["id"], c["name"], c["tier"], round(c["irn"],3), round(c["perf"],3)] + [round(c[a["id"]],3) for a in apps])
    ws.column_dimensions["B"].width = 40

    # 07_Sources
    ws = wb.create_sheet("07_Sources")
    header(ws, 1, ["Bloc", "Source", "Fraîcheur", "URL"])
    for row in SOURCES:
        ws.append(row)
    ws.column_dimensions["B"].width = 40; ws.column_dimensions["D"].width = 60

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


SOURCES = [
    ["Performance", "OpenLM Chatbot Arena", "juin 2026", "https://openlm.ai/chatbot-arena/"],
    ["Performance", "Artificial Analysis — leaderboards", "juin 2026", "https://artificialanalysis.ai/leaderboards/models"],
    ["Performance", "LMArena", "juin 2026", "https://lmarena.ai/leaderboard"],
    ["Modèle", "Mistral Large 3 (open-weight, Apache 2.0)", "déc. 2025", "https://mistral.ai/news/mistral-3/"],
    ["Modèle", "Mistral Medium 3.5", "2026", "https://docs.mistral.ai/resources/changelogs"],
    ["Modèle", "Mistral Devstral 2 / Vibe", "2026", "https://mistral.ai/news/devstral-2-vibe-cli/"],
    ["Modèle", "Qwen3.6-27B Coder (Apache 2.0)", "avr. 2026", "https://qwen.ai/blog?id=qwen3.6-27b"],
    ["Modèle", "DeepSeek V4", "2026", "https://api-docs.deepseek.com/"],
    ["Modèle", "Claude Opus 4.8", "juin 2026", "https://artificialanalysis.ai/articles/claude-opus-4-8-analysis-and-benchmarks"],
    ["Modèle", "OpenAI GPT-5.5", "2026", "https://artificialanalysis.ai/articles/openai-gpt5-5-is-the-new-leading-AI-model"],
    ["Souverain", "Albert API — Etalab / DINUM (SecNumCloud)", "2026", "https://www.numerique.gouv.fr/offre-accompagnement/expertise-albert-ia-etat/"],
    ["IRN", "Référentiel aDRI", "2026", "https://gitlab.com/digitalresilienceinitiative/adri-irn"],
    ["Coût GPU", "RunPod / CoreWeave public pricing", "2026", "https://www.runpod.io/pricing"],
]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", default=DEFAULT_SOURCE, type=Path)
    ap.add_argument("--json",   default=DEFAULT_JSON,   type=Path)
    ap.add_argument("--xlsx",   default=DEFAULT_XLSX,   type=Path)
    ap.add_argument("--no-xlsx", action="store_true")
    args = ap.parse_args()

    source = json.loads(Path(args.source).read_text(encoding="utf-8"))
    data = build(source)

    args.json.parent.mkdir(parents=True, exist_ok=True)
    args.json.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✅ JSON : {args.json} ({len(data['configurations'])} configs, {len(data['app_profiles'])} profils)")

    if not args.no_xlsx:
        try:
            write_xlsx(data, source, args.xlsx)
            print(f"✅ XLSX : {args.xlsx}")
        except ImportError:
            print("⚠ openpyxl absent — XLSX non régénéré (pip install openpyxl)")


if __name__ == "__main__":
    main()
